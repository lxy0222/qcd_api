from openpyxl import load_workbook
from  common.read_path import *

class DoInitData:
    '''读取初始数据，如电话号码'''
    @classmethod
    def read_initdata(cls,i,j):
        '''i=行数
        j=列数'''
        wb = load_workbook(testcase_path)
        sheet = wb['telnumber']
        inittel=sheet.cell(i,j).value
        return inittel
    '''更新初始数据，如未注册号码'''
    def updata_initdata(self,i,j,newdata):
        '''i=行数
               j=列数'''
        wb = load_workbook(testcase_path)
        sheet = wb['telnumber']
        sheet.cell(i,j).value=newdata
        wb.save(testcase_path)