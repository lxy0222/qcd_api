from openpyxl import load_workbook
from common.read_config import ReadConfig
from common import read_path
from common.get_data import GetData
from common.do_initdata import DoInitData
class DoExcel:
    '''获取测试用例表格表头'''
    @classmethod
    def get_title(cls,wbname,sheetname):
        wb=load_workbook(wbname)
        sheet=wb[sheetname]
        title=[]
        for i in range(1,sheet.max_column+1): #遍历表格每一列
            titles=sheet.cell(1,i).value
            title.append(titles)
        return title

    '''获取测试用例数据'''
    @classmethod
    def get_data(cls,wbname):
        mode = eval(ReadConfig().read_config(read_path.config_path, 'MODE', 'mode')) #读取配置文件，返回为一个字典
        wb = load_workbook(wbname)
        test_data=[] #存放收集的每一条用例组合而成的字典
        for key in mode: #遍历mode里的key值
            sheet = wb[key]
            titles = DoExcel().get_title(wbname,key)
            #mode等于all时读取该sheet下所有用例
            if mode[key]=='all':
                for j in range(2,sheet.max_row+1):  #遍历除表头行的其他行
                    test_lists={}  #存放取出的表格数据组合的key，value值
                    for i in range(1,sheet.max_column+1):   #遍历每一行
                        #替换有noreg_tel的参数
                        if str(sheet.cell(j,i).value).find('${noreg_tel}')!=-1:
                            tel=getattr(GetData,'noreg_tel')  #利用反射获取初始值
                            #用初始值替换
                            test_lists[titles[i-1]]=str(sheet.cell(j,i).value).replace('${noreg_tel}',str(tel))
                            #更新未注册号码
                            DoInitData().updata_initdata(1,1,int(tel)+2)
                        # 替换有noreg_tel1的参数
                        elif str(sheet.cell(j,i).value).find('${noreg_tel1}')!=-1:
                            tel=int(getattr(GetData,'noreg_tel'))+1
                            test_lists[titles[i-1]] = str(sheet.cell(j,i).value).replace('${noreg_tel1}',str(tel))
                        # 替换有login_tel的参数
                        elif str(sheet.cell(j,i).value).find('${login_tel}')!=-1:
                            tel=getattr(GetData,'login_tel')
                            test_lists[titles[i-1]] = str(sheet.cell(j,i).value).replace('${login_tel}',str(tel))
                        # 替换有admin_tel的参数
                        elif str(sheet.cell(j, i).value).find('${admin_tel}') != -1:
                            tel = getattr(GetData, 'admin_tel')
                            test_lists[titles[i - 1]] = str(sheet.cell(j, i).value).replace('${admin_tel}', str(tel))
                        # 替换有loan_member_id的参数
                        elif str(sheet.cell(j, i).value).find('${loan_member_id}') != -1:
                            id = getattr(GetData, 'loan_member_id')
                            test_lists[titles[i - 1]] = str(sheet.cell(j, i).value).replace('${loan_member_id}', str(id))
                        # 替换有memberID的参数
                        elif str(sheet.cell(j, i).value).find('${memberID}') != -1:
                            id = getattr(GetData, 'memberID')
                            test_lists[titles[i - 1]] = str(sheet.cell(j, i).value).replace('${memberID}', str(id))
                        else:
                            test_lists[titles[i-1]] = sheet.cell(j,i).value
                    test_data.append(test_lists)
            #遍历某个sheet下某几条用例
            else:
                for case_id in mode[key]:  #遍历mode里的value值
                    test_lists={}
                    for i in range(1,sheet.max_column+1):  #遍历列
                        if str(sheet.cell(case_id+1,i).value).find('${noreg_tel}')!=-1:
                            tel=getattr(GetData,'noreg_tel')
                            test_lists[titles[i-1]]=str(sheet.cell(case_id+1,i).value).replace('${noreg_tel}',str(tel))   #组合字典
                            DoInitData().updata_initdata(1,1,int(tel)+2)
                        elif str(sheet.cell(case_id+1, i).value).find('${noreg_tel1}') != -1:
                            tel = int(getattr(GetData, 'noreg_tel')) + 1
                            test_lists[titles[i-1]] = str(sheet.cell(case_id+1, i).value).replace('${noreg_tel1}', str(tel))
                        elif str(sheet.cell(case_id+1,i).value).find('${login_tel}')!=-1:
                            tel=getattr(GetData,'login_tel')
                            test_lists[titles[i-1]] = str(sheet.cell(case_id+1,i).value).replace('${login_tel}',str(tel))
                        elif str(sheet.cell(case_id+1,i).value).find('${loan_member_id}')!=-1:
                            id=getattr(GetData,'loan_member_id')
                            test_lists[titles[i-1]] = str(sheet.cell(case_id+1,i).value).replace('${loan_member_id}',str(id))
                        elif str(sheet.cell(case_id+1,i).value).find('${admin_tel}')!=-1:
                            tel=getattr(GetData,'admin_tel')
                            test_lists[titles[i-1]] = str(sheet.cell(case_id+1,i).value).replace('${admin_tel}',str(tel))
                        elif str(sheet.cell(case_id, i).value).find('${memberID}') != -1:
                            id = getattr(GetData, 'memberID')
                            test_lists[titles[i - 1]] = str(sheet.cell(case_id, i).value).replace('${memberID}', str(id))
                        else:
                            test_lists[titles[i-1]]=sheet.cell(case_id+1,i).value
                    test_data.append(test_lists)
        return test_data #返回组装的测试用例
    '''写回用例执行结果'''
    @classmethod
    def write_back(cls,wbname,sheetname,case_id,json_result,passorfail):
        wb = load_workbook(wbname)
        sheet = wb[sheetname]
        sheet.cell(case_id+1,8).value=json_result
        sheet.cell(case_id+1,9).value=passorfail
        wb.save(wbname)
    '''写回sql执行结果'''
    @classmethod
    def write_checkres(cls,wbname,sheetname,case_id,check_res):
        wb = load_workbook(wbname)
        sheet = wb[sheetname]
        sheet.cell(case_id + 1,11).value = check_res
        wb.save(wbname)
if __name__ == '__main__':
    from common.read_path import *
    test_data=DoExcel.get_data(testcase_path)
    print(test_data)
    DoExcel.write_checkres(testcase_path,'recharge',2,'shiabi')